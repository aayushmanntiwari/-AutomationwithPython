import openpyxl as xl
# Add function
def Add():
    wb = xl.load_workbook('transactions.xlsx')  # write the spreedsheet(.xlsx) file name here you to Update
    sheet = wb['Sheet1']
    y = int(input("Enter the targeted row where value start: "))
    z = int(input("Enter the target column where these value belong: "))
    m = int(input("Enter where you want to import this Updated Value: "))
    for row in range(y, sheet.max_row + 1):
        x = int(input(f"Enter the value you want to Add for Row {row} with the origianl value: "))
        cell = sheet.cell(row, z)
        print(type(cell))
        corrected_cell_price = cell.value + x
        print(type(corrected_cell_price))
        print(corrected_cell_price)
        corrected_cell = sheet.cell(row,m)
        corrected_cell.value = corrected_cell_price
    wb.save('transactions.xlsx')  # You can save the file formate with any name