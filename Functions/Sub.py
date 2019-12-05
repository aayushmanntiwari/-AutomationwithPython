import openpyxl as xl
def Sub():
    wb = xl.load_workbook('')    # write the spreedsheet(.xlsx) file name here you to Update
    sheet = wb['Sheet1']
    x = int(input("Enter the targeted row where value start: "))
    z = int(input("Enter the target column where these value belong: "))
    m = int(input("Enter where you want to import this Updated Value: "))
    for rows in range(x,sheet.max_row + 1):
        y = int(input(f"Enter the value you want to Substract for Row {rows} with the origianl value: "))
        cell = sheet.cell(rows,z)
        corrected_cell_pair = cell.value + y
        corrected_cell = sheet.cell(rows,m)
        corrected_cell_pair.value = corrected_cell
    wb.save()   # You can save the file formate with any name