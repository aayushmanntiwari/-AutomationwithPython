import openpyxl as xl
def Div():
    wb = xl.load_workbook('')  # write the spreedsheet(.xlsx) file name here you to Update
    sheet = wb['Sheet1']
    x = int(input("Enter the targeted row: "))
    y = int(input("Enter the target column: "))
    m = int(input("Enter where you want to import this Updated Value: "))
    for rows in range(x, sheet.max_row + 1):
        z = int(input(f"Enter the value you want to Divide for Row {rows} with the origianl value: "))
        cell = sheet.cell(rows, y)
        corrected_cell_pair = cell.value / z
        corrected_cell = sheet.cell(rows,m)
        corrected_cell_pair.value = corrected_cell

    wb.save()  # You can save the file formate with any name